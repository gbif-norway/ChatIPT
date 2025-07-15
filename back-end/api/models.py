import traceback
from django.db import models
from django.utils.translation import gettext_lazy as _
from django.contrib.postgres.fields import ArrayField
from django.contrib.auth.models import AbstractUser
from api import agent_tools
from api.helpers.openai_helpers import create_chat_completion
from picklefield.fields import PickledObjectField
import pandas as pd
from django.template.loader import render_to_string
import os
import json
import ujson
import logging
import openpyxl
import tempfile
import re
import numpy as np
import io


class CustomUser(AbstractUser):
    """Custom user model with ORCID integration"""
    orcid_id = models.CharField(max_length=50, blank=True, help_text="ORCID identifier")
    orcid_access_token = models.TextField(blank=True, help_text="ORCID OAuth access token")
    orcid_refresh_token = models.TextField(blank=True, help_text="ORCID OAuth refresh token")
    institution = models.CharField(max_length=500, blank=True, help_text="User's institution")
    department = models.CharField(max_length=500, blank=True, help_text="User's department")
    country = models.CharField(max_length=100, blank=True, help_text="User's country")
    
    def __str__(self):
        return f"{self.email} ({self.orcid_id})"
    
    class Meta:
        verbose_name = "User"
        verbose_name_plural = "Users"


class Dataset(models.Model):
    created_at = models.DateTimeField(auto_now_add=True)
    user = models.ForeignKey(CustomUser, on_delete=models.CASCADE, related_name='datasets', null=True, blank=True)
    orcid = models.CharField(max_length=2000, blank=True)
    file = models.FileField(upload_to='user_files')
    title = models.CharField(max_length=2000, blank=True, default='')
    structure_notes = models.TextField(default='', blank=True)
    description = models.CharField(max_length=2000, blank=True, default='')
    published_at = models.DateTimeField(null=True, blank=True)
    rejected_at = models.DateTimeField(null=True, blank=True)
    dwca_url = models.CharField(max_length=2000, blank=True)
    gbif_url = models.CharField(max_length=2000, blank=True)
    user_language = models.CharField(max_length=100, blank=True)

    class DWCCore(models.TextChoices):
        EVENT = 'event_occurrences'
        OCCURRENCE = 'occurrence'
        TAXONOMY = 'taxonomy'
    dwc_core = models.CharField(max_length=30, choices=DWCCore.choices, blank=True)

    class DWCExtensions(models.TextChoices):
        SIMPLE_MULTIMEDIA = 'simple_multimedia'
        MEASUREMENT_OR_FACT = 'measurement_or_fact'
        GBIF_RELEVE = 'gbif_releve'
    dwc_extensions = ArrayField(base_field=models.CharField(max_length=500, choices=DWCExtensions.choices), null=True, blank=True)

    @property
    def filename(self):
        return os.path.basename(self.file.name)

    def next_agent(self):
        self.refresh_from_db()
        if self.rejected_at:
            logger = logging.getLogger(__name__)
            logger.info('rejected')
            return None

        next_agent = self.agent_set.filter(completed_at=None).first()
        if next_agent:
            return next_agent

        last_completed_agent = self.agent_set.last()  # self.agent_set.exclude(completed_at=None).last()
        logger = logging.getLogger(__name__)
        logger.info(f'No next agent found, making new agent for new task based on {last_completed_agent}')
        if last_completed_agent:
            next_task = Task.objects.filter(id__gt=last_completed_agent.task.id).first()
            if next_task:
                next_task.create_agent_with_system_messages(dataset=self)
                return self.next_agent()
            else:
                logger.info(f'PUBLISHED {self.published_at}')
                return None  # It's been published... self.published_at = datetime.now() # self.save()
        else:
            # No agents exist for this dataset - create the first agent
            logger.info('No agents found for dataset, creating first agent')
            first_task = Task.objects.first()
            if not first_task:
                raise Exception('No tasks are configured in the system. Please contact the administrator to load the required tasks.')
            
            # Get tables for this dataset
            tables = Table.objects.filter(dataset=self)
            if not tables.exists():
                raise Exception('No tables found for this dataset. Please ensure the dataset has been properly processed.')
            
            first_task.create_agent_with_system_messages(dataset=self)
            return self.next_agent()

    @staticmethod
    def get_dfs_from_user_file(file, file_name):
        try:
            file_content = file.read()
            file_io = io.StringIO(file_content.decode('utf-8', errors='surrogateescape'))
            df = pd.read_csv(file_io, dtype='str', encoding='utf-8', encoding_errors='surrogateescape', sep=None, engine='python', header=0)
            return {file_name: df}
        except Exception as e:
            try:
                workbook = openpyxl.load_workbook(file)
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.data_type == 'f':  # 'f' indicates a formula
                                cell.value = '' # f'[FORMULA: {cell.value}]'
                    for merged_cell in list(sheet.merged_cells.ranges):
                        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
                        value = sheet.cell(row=min_row, column=min_col).value
                        sheet.unmerge_cells(str(merged_cell))
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                sheet.cell(row=row, column=col).value = f"{value} [UNMERGED CELL]"
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    temp_file_name = tmp.name
                    workbook.save(temp_file_name)
                dfs = pd.read_excel(temp_file_name, dtype='str', sheet_name=None)
                os.remove(temp_file_name)
                return dfs
            except ValueError as ve:
                return {"error": f"Unable to read workbook: {str(ve)}. The file may contain invalid XML or be corrupted."}
            except Exception as e:
                return {"error": f"An error occurred while processing the file: {str(e)}."}

    class Meta:
        get_latest_by = 'created_at'
        ordering = ['created_at']


class Task(models.Model):  # See tasks.yaml for the only objects this model is populated with
    name = models.CharField(max_length=300, unique=True)
    text = models.TextField()

    class Meta:
        get_latest_by = 'id'
        ordering = ['id']

    @property
    def functions(self):
        functions = [agent_tools.SetBasicMetadata.__name__,
                    agent_tools.SetAgentTaskToComplete.__name__,
                    agent_tools.Python.__name__,
                    agent_tools.BasicValidationForSomeDwCTerms.__name__,
                    agent_tools.RollBack.__name__,
                    agent_tools.PublishDwC.__name__]
        return [getattr(agent_tools, f) for f in functions]

    def create_agent_with_system_messages(self, dataset:Dataset):
        tables = Table.objects.filter(dataset=dataset)
        return Agent.create_with_system_message(dataset=dataset, task=self, tables=tables)


class Table(models.Model):
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)
    dataset = models.ForeignKey(Dataset, on_delete=models.CASCADE)
    title = models.CharField(max_length=200, blank=True)
    df = PickledObjectField()
    description = models.CharField(max_length=2000, blank=True)

    @property
    def df_json(self):
        df = self.make_columns_unique(self.df)
        # return df.to_json(orient='records', date_format='iso')
        df = df.replace([np.inf, -np.inf], np.nan)
        df = df.where(pd.notnull(df), None)
        def clean_strings_in_df(df):
            for col in df.select_dtypes(include=['object']).columns:
                df[col] = df[col].astype(str).apply(
                    lambda x: x.encode('utf-8', 'replace').decode('utf-8')
                )
            return df
        df = clean_strings_in_df(df)
        try:
            return df.to_json(orient='records', date_format='iso', force_ascii=False)
        except Exception as e:
            raise Exception(f"Serialization failed after cleaning data: {e}")

    def _snapshot_df(self, df_obj):
        max_rows, max_columns, max_str_len = 10, 10, 70

        # Truncate long strings in cells
        df = df_obj.apply(lambda col: col.astype(str).map(lambda x: (x[:max_str_len - 3] + '...') if len(x) > max_str_len else x))

        # Truncate columns
        if len(df.columns) > max_columns:
            left = df.iloc[:, :max_columns//2]
            right = df.iloc[:, -max_columns//2:]
            middle = pd.DataFrame({ '...': ['...']*len(df) }, index=df.index)
            df = pd.concat([left, middle, right], axis=1)

        df.fillna('', inplace=True)

        # Truncate rows
        if len(df) > max_rows:
            top = df.head(max_rows // 2)
            bottom = df.tail(max_rows // 2)
            middle = pd.DataFrame({col: ['...'] for col in df.columns}, index=[0])  # Use a temporary numeric index for middle
            df = pd.concat([top, middle, bottom], ignore_index=True)
            # df = '\n'.join([top, middle, bottom])

        return df

    @property
    def str_snapshot(self):
        df = self.make_columns_unique(self.df)
        original_rows, original_cols = self.df.shape
        return self._snapshot_df(df).to_string() + f"\n\n[{original_rows} rows x {original_cols} columns]"

    def make_columns_unique(self, df):
        cols = pd.Series(df.columns)
        nan_count = 0
        for i, col in enumerate(cols):
            if pd.isna(col):
                nan_count += 1
                cols[i] = f"NaN ({nan_count})"
            elif (cols == col).sum() > 1:
                dup_indices = cols[cols == col].index
                for j, idx in enumerate(dup_indices, start=1):
                    if j > 1:
                        cols[idx] = f"{col} ({j})"

        df.columns = cols
        return df


class Agent(models.Model):
    created_at = models.DateTimeField(auto_now_add=True)
    completed_at = models.DateTimeField(null=True, blank=True)
    dataset = models.ForeignKey(Dataset, on_delete=models.CASCADE)
    task = models.ForeignKey(Task, on_delete=models.CASCADE)
    tables = models.ManyToManyField(Table, blank=True)
    busy_thinking = models.BooleanField(default=False)

    class Meta:
        get_latest_by = 'created_at'
        ordering = ['created_at']

    @classmethod
    def create_with_system_message(cls, dataset, task, tables):
        if not task:
            raise ValueError("Task cannot be None. Please ensure tasks are loaded in the database.")
        
        agent = cls.objects.create(dataset=dataset, task=task)
        agent.tables.set([t.id for t in tables])
        system_message_text = render_to_string('prompt.txt', context={ 'agent': agent, 'all_tasks_count': Task.objects.all().count() })
        logger = logging.getLogger(__name__)
        logger.info(system_message_text)
        # import pdb; pdb.set_trace()
        Message.objects.create(agent=agent, openai_obj={'content': system_message_text, 'role': Message.Role.SYSTEM})

    def next_message(self):
        last_message = self.message_set.last()
        print(f'Last message role: {last_message.role}, Completed at value for this agent: {self.completed_at}')
        if last_message.role == Message.Role.ASSISTANT or self.completed_at:
            return None
        if self.busy_thinking:
            return last_message

        # Otherwise we need to send it to GPT, last message was from the user, was the return from a function, or was the starting system message
        self.busy_thinking = True
        self.save()
        try:
            response_message = create_chat_completion(self.message_set.all(), self.task.functions)
        except Exception as e:
            error_message = f'Unfortunately there was a problem querying the OpenAI API. Try again later, and please report this error to the developers. Full error: {e}'
            print(e)
            print("Error in next_message:")
            traceback.print_exc()
            self.busy_thinking = False
            self.save()
            return [Message.objects.create(agent=self, openai_obj={'role': Message.Role.ASSISTANT, 'content': error_message})]

        message = Message.objects.create(agent=self, openai_obj=response_message.dict())  # response_message.__dict__
        if not response_message.tool_calls:  # It's a simple assistant message
            self.busy_thinking = False
            self.save()
            return [message]

        messages = [message]  # Store the API response which requests the tool calls
        for tool_call in response_message.tool_calls:  # Occasionally a single API response requests multiple tool calls
            try:
                result = self.run_function(tool_call.function)
            except Exception as e:
                result = f'ERROR CALLING FUNCTION: Invalid JSON or code provided in your last response (Calling {tool_call.function.name} with the given arguments for {tool_call.id}), please try again. \nError: {e}'

            messages.append(Message.create_function_message(agent=self, function_result=result, tool_call_id=tool_call.id))

        self.refresh_from_db()  # Necessary so that completed_at doesn't get overwritten
        self.busy_thinking = False
        self.save()
        return messages

    def run_function(self, fn):
        function_model_class = getattr(agent_tools, fn.name)
        fnargs = fn.arguments
        if fn.name == 'Python':
            if not re.sub(r'[\s"\']', '', fn.arguments).startswith('{code'):
                fnargs = json.dumps({'code': fn.arguments})
        fn_args = json.loads(fnargs, strict=False)
        function_model_obj = function_model_class(**fn_args)
        return function_model_obj.run()


class Message(models.Model):
    created_at = models.DateTimeField(auto_now_add=True)
    agent = models.ForeignKey(Agent, on_delete=models.CASCADE)
    openai_obj = models.JSONField(null=True, blank=True)

    class Role(models.TextChoices):
        USER = 'user'
        SYSTEM = 'system'
        ASSISTANT = 'assistant'
        TOOL = 'tool'

    @classmethod
    def create_function_message(cls, agent, function_result, tool_call_id):
        return cls.objects.create(agent=agent, openai_obj={'content': function_result, 'role': cls.Role.TOOL, 'tool_call_id': tool_call_id})

    @property
    def role(self):
        return self.Role(self.openai_obj['role'])

    class Meta:
        get_latest_by = 'created_at'
        ordering = ['created_at']

